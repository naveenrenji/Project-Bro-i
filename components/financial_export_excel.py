"""
Excel workbook export for the Program Financial Estimation.

Generates a presentation-ready, multi-sheet .xlsx file with:
  1. Executive Summary  (KPIs + P&L table + charts)
  2. Cohort Matrix      (enrollment heatmap + area chart)
  3. Revenue            (per-term detail + bar chart)
  4. Costs              (per-term breakdown + stacked bar chart)
  5. Assumptions        (all input parameters)

All styling aligns with Stevens CPE Brand Guidelines (Oct 2025).
"""

from __future__ import annotations
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle, numbers,
)
from openpyxl.chart import BarChart, LineChart, AreaChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import pandas as pd

# ── Brand colours ──────────────────────────────────────────────────────
_RED_HEX   = "A32638"
_DGRAY_HEX = "363D45"
_GRAY_HEX  = "7F7F7F"
_LGRAY_HEX = "E4E5E6"
_WHITE_HEX = "FFFFFF"
_GREEN_HEX = "2E7D32"

# ── Reusable styles ───────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor=_RED_HEX)
_HEADER_FONT = Font(name="Calibri", bold=True, color=_WHITE_HEX, size=11)
_BODY_FONT   = Font(name="Calibri", size=10, color=_DGRAY_HEX)
_BOLD_FONT   = Font(name="Calibri", size=10, bold=True, color=_DGRAY_HEX)
_TITLE_FONT  = Font(name="Calibri", size=16, bold=True, color=_RED_HEX)
_SUB_FONT    = Font(name="Calibri", size=12, bold=True, color=_DGRAY_HEX)
_KPI_FONT    = Font(name="Calibri", size=20, bold=True, color=_DGRAY_HEX)
_KPI_LABEL   = Font(name="Calibri", size=9, color=_GRAY_HEX)
_ALT_FILL    = PatternFill("solid", fgColor=_LGRAY_HEX)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color=_LGRAY_HEX),
)
_MONEY_FMT   = '$#,##0'
_MONEY_NEG   = '$#,##0;[Red]($#,##0)'
_PCT_FMT     = '0.0%'

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
_ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")


# ── Utility ───────────────────────────────────────────────────────────

def _set_col_widths(ws, widths: dict[str, float]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_header_row(ws, row: int, values: list[str], start_col: int = 1):
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER


def _write_data_row(ws, row: int, values: list, start_col: int = 1,
                    bold: bool = False, fmt: str | None = None, alt: bool = False):
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font = _BOLD_FONT if bold else _BODY_FONT
        cell.alignment = _ALIGN_RIGHT if isinstance(val, (int, float)) else _ALIGN_LEFT
        cell.border = _THIN_BORDER
        if alt:
            cell.fill = _ALT_FILL
        if fmt and isinstance(val, (int, float)):
            cell.number_format = fmt


def _auto_width(ws, min_width: float = 10, max_width: float = 22):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# ═══════════════════════════ SHEET BUILDERS ═══════════════════════════

def _build_executive_summary(wb: Workbook, inputs: dict, results: dict):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = _RED_HEX

    # Title block
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{inputs['program_name']} - Financial Estimation"
    c.font = _TITLE_FONT
    c.alignment = _ALIGN_LEFT

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    start = inputs.get("start_fy", 2026)
    c.value = f"{inputs['projection_years']}-Year Outlook  |  FY{start}-FY{start + inputs['projection_years'] - 1}  |  {inputs['delivery_format']} delivery"
    c.font = Font(name="Calibri", size=10, italic=True, color=_GRAY_HEX)

    # KPI row (row 4)
    kpis = [
        ("Total Revenue", results["total_revenue"], _MONEY_FMT),
        ("Total Cost",    results["total_cost"],    _MONEY_FMT),
        ("Net P&L",       results["total_net"],     _MONEY_NEG),
        ("Break-Even",    results["break_even_year"] or "N/A", None),
    ]
    for i, (label, val, fmt) in enumerate(kpis):
        col = 1 + i * 2
        lbl_cell = ws.cell(row=4, column=col, value=label)
        lbl_cell.font = _KPI_LABEL
        val_cell = ws.cell(row=5, column=col, value=val)
        val_cell.font = _KPI_FONT
        if fmt and isinstance(val, (int, float)):
            val_cell.number_format = fmt

    # P&L table
    pl = results["pl_summary"]
    header_row = 7
    cols = list(pl.columns)
    _write_header_row(ws, header_row, cols)

    for r_idx, (_, row) in enumerate(pl.iterrows()):
        is_total = row["Fiscal Year"] == "Total"
        is_alt = r_idx % 2 == 1 and not is_total
        vals = []
        for col in cols:
            v = row[col]
            vals.append(v)
        data_row = header_row + 1 + r_idx
        _write_data_row(ws, data_row, vals, bold=is_total, alt=is_alt)
        # Apply formats
        for ci, col_name in enumerate(cols):
            cell = ws.cell(row=data_row, column=1 + ci)
            if col_name in ("Revenue", "Cost", "Net", "Cumulative"):
                cell.number_format = _MONEY_NEG
            elif col_name == "Net Margin %":
                cell.number_format = '0.0"%"'

    last_data_row = header_row + len(pl)

    # Bar chart: Revenue vs Cost
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Revenue vs Cost by Fiscal Year"
    chart.y_axis.numFmt = _MONEY_FMT
    chart.style = 10
    chart.width = 18
    chart.height = 12

    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_data_row - 1)
    rev_ref = Reference(ws, min_col=2, min_row=header_row, max_row=last_data_row - 1)
    cost_ref = Reference(ws, min_col=3, min_row=header_row, max_row=last_data_row - 1)
    chart.add_data(rev_ref, titles_from_data=True)
    chart.add_data(cost_ref, titles_from_data=True)
    chart.set_categories(cats)
    if len(chart.series) >= 1:
        chart.series[0].graphicalProperties.solidFill = _RED_HEX
    if len(chart.series) >= 2:
        chart.series[1].graphicalProperties.solidFill = _DGRAY_HEX
    ws.add_chart(chart, f"A{last_data_row + 2}")

    # Line chart: Cumulative Net
    line = LineChart()
    line.title = "Cumulative Net P&L"
    line.y_axis.numFmt = _MONEY_FMT
    line.style = 10
    line.width = 18
    line.height = 12
    cum_ref = Reference(ws, min_col=5, min_row=header_row, max_row=last_data_row - 1)
    line.add_data(cum_ref, titles_from_data=True)
    line.set_categories(cats)
    if line.series:
        line.series[0].graphicalProperties.line.solidFill = _RED_HEX
    ws.add_chart(line, f"A{last_data_row + 18}")

    _auto_width(ws)
    ws.freeze_panes = f"A{header_row + 1}"


def _build_cohort_matrix(wb: Workbook, inputs: dict, results: dict):
    ws = wb.create_sheet("Cohort Matrix")
    ws.sheet_properties.tabColor = _DGRAY_HEX

    cm = results["cohort_matrix"]
    headers = ["Cohort"] + list(cm.columns)
    _write_header_row(ws, 1, headers)

    for r_idx, (idx_label, row) in enumerate(cm.iterrows()):
        vals = [idx_label] + [round(v, 1) for v in row.values]
        _write_data_row(ws, 2 + r_idx, vals, alt=r_idx % 2 == 1, fmt='#,##0.0')

    last_row = 1 + len(cm)

    # Sum row
    sum_vals = ["Sum"]
    for col in cm.columns:
        sum_vals.append(round(cm[col].sum(), 1))
    _write_data_row(ws, last_row + 1, sum_vals, bold=True, fmt='#,##0.0')

    # Area chart: total active per term
    labels = results["term_labels"]
    active = results["total_active"]

    chart_start = last_row + 3
    ws.cell(row=chart_start, column=1, value="Term").font = _HEADER_FONT
    ws.cell(row=chart_start, column=2, value="Total Active").font = _HEADER_FONT
    for i, (lbl, val) in enumerate(zip(labels, active)):
        ws.cell(row=chart_start + 1 + i, column=1, value=lbl)
        ws.cell(row=chart_start + 1 + i, column=2, value=round(val, 1))

    chart = AreaChart()
    chart.title = "Total Active Students per Term"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    cats = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + len(labels))
    data = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + len(labels))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = _RED_HEX
    ws.add_chart(chart, f"D{last_row + 3}")

    _auto_width(ws)
    ws.freeze_panes = "B2"


def _build_revenue(wb: Workbook, inputs: dict, results: dict):
    ws = wb.create_sheet("Revenue")
    ws.sheet_properties.tabColor = _RED_HEX

    rev = results["revenue_df"]
    headers = list(rev.columns)
    _write_header_row(ws, 1, headers)

    for r_idx, (_, row) in enumerate(rev.iterrows()):
        vals = list(row.values)
        _write_data_row(ws, 2 + r_idx, vals, alt=r_idx % 2 == 1)
        for ci, col in enumerate(headers):
            cell = ws.cell(row=2 + r_idx, column=1 + ci)
            if col in ("Base Revenue", "Revenue", "Tuition/Credit"):
                cell.number_format = _MONEY_FMT
            elif col == "Active Students":
                cell.number_format = '#,##0.0'

    last_row = 1 + len(rev)

    # Sum row
    sum_vals = ["Total", "", "", "",
                round(rev["Base Revenue"].sum(), 2),
                round(rev["Revenue"].sum(), 2)]
    _write_data_row(ws, last_row + 1, sum_vals, bold=True, fmt=_MONEY_FMT)

    # Bar chart
    chart = BarChart()
    chart.title = "Revenue per Term"
    chart.y_axis.numFmt = _MONEY_FMT
    chart.style = 10
    chart.width = 20
    chart.height = 12
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    data = Reference(ws, min_col=6, min_row=1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = _RED_HEX
    ws.add_chart(chart, f"A{last_row + 3}")

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_costs(wb: Workbook, inputs: dict, results: dict):
    ws = wb.create_sheet("Costs")
    ws.sheet_properties.tabColor = _DGRAY_HEX

    cost = results["cost_df"]
    headers = list(cost.columns)
    _write_header_row(ws, 1, headers)

    for r_idx, (_, row) in enumerate(cost.iterrows()):
        vals = list(row.values)
        _write_data_row(ws, 2 + r_idx, vals, alt=r_idx % 2 == 1)
        for ci, col in enumerate(headers):
            cell = ws.cell(row=2 + r_idx, column=1 + ci)
            if col != "Term":
                cell.number_format = _MONEY_FMT

    last_row = 1 + len(cost)

    # Sum row
    sum_vals = ["Total"]
    for col in headers[1:]:
        sum_vals.append(round(cost[col].sum(), 2))
    _write_data_row(ws, last_row + 1, sum_vals, bold=True, fmt=_MONEY_FMT)

    # Stacked bar chart
    cost_cols = ["Faculty", "TA", "Course Dev", "Variable OH", "Fixed OH", "CAC"]
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Cost Components by Term"
    chart.y_axis.numFmt = _MONEY_FMT
    chart.style = 10
    chart.width = 22
    chart.height = 13

    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    colours = [_RED_HEX, _DGRAY_HEX, _GRAY_HEX, _LGRAY_HEX, "5A6577", _GREEN_HEX]
    for i, col_name in enumerate(cost_cols):
        col_idx = headers.index(col_name) + 1
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for i, s in enumerate(chart.series):
        s.graphicalProperties.solidFill = colours[i % len(colours)]
    ws.add_chart(chart, f"A{last_row + 3}")

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_assumptions(wb: Workbook, inputs: dict, results: dict):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = _GRAY_HEX

    _write_header_row(ws, 1, ["Parameter", "Value"])

    sections: list[tuple[str, list[tuple[str, Any, str | None]]]] = [
        ("Program Structure", [
            ("Program Name",        inputs["program_name"], None),
            ("Total Courses",       inputs["total_courses"], None),
            ("Credits per Course",  inputs["credits_per_course"], None),
            ("Delivery Format",     inputs["delivery_format"], None),
            ("Include Summer",      "Yes" if inputs.get("include_summer") else "No", None),
            ("Projection Years",    inputs["projection_years"], None),
            ("Starting FY",         inputs.get("start_fy", 2026), None),
        ]),
        ("Course Development", [
            ("New Courses to Develop",   inputs["courses_to_develop"], None),
            ("Dev Cost per Course",      inputs["dev_cost_per_course"], _MONEY_FMT),
            ("Courses to Revise",        inputs["courses_to_revise"], None),
            ("Revision Cost %",          inputs["revision_cost_pct"], _PCT_FMT),
            ("Amortisation Terms",       inputs["dev_amortization_terms"], None),
        ]),
        ("Enrollment & Growth", [
            ("Initial Fall-1 Intake",  inputs["initial_intake"], None),
            ("Fall Growth %",          inputs["fall_growth_rate"], _PCT_FMT),
            ("Spring Growth %",        inputs["spring_growth_rate"], _PCT_FMT),
            ("Summer Growth %",        inputs.get("summer_growth_rate", 0), _PCT_FMT),
        ]),
        ("Retention", [
            ("Early Retention %",        inputs["early_retention_rate"], _PCT_FMT),
            ("Late Retention %",         inputs["late_retention_rate"], _PCT_FMT),
            ("Threshold Term",           inputs["retention_threshold_term"], None),
        ]),
        ("Tuition & Revenue", [
            ("Tuition per Credit",       inputs["tuition_per_credit"], _MONEY_FMT),
            ("Credits per Term",         inputs["credits_per_term"], None),
            ("Tuition Inflation %/yr",   inputs["tuition_inflation_pct"], _PCT_FMT),
        ]),
        ("Faculty & TA", [
            ("Faculty $/Section/Term",   inputs["faculty_cost_per_section"], _MONEY_FMT),
            ("Sections per Term",        inputs["sections_per_term"], None),
            ("TA:Student Ratio",         inputs["ta_student_ratio"], None),
            ("TA Hourly Rate",           inputs["ta_hourly_rate"], _MONEY_FMT),
            ("TA Hours/Week",            inputs["ta_hours_per_week"], None),
            ("Weeks per Session",        inputs["weeks_per_session"], None),
        ]),
        ("Overhead & Marketing", [
            ("Variable OH/Student",      inputs["variable_overhead_per_student"], _MONEY_FMT),
            ("Fixed OH/Term",            inputs["fixed_overhead_per_term"], _MONEY_FMT),
            ("CAC/Student",              inputs["cac_per_student"], _MONEY_FMT),
            ("Cost Inflation %/yr",      inputs["cost_inflation_pct"], _PCT_FMT),
        ]),
    ]

    row = 2
    for section_name, params in sections:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=section_name)
        cell.font = _SUB_FONT
        cell.fill = PatternFill("solid", fgColor=_LGRAY_HEX)
        row += 1

        for label, value, fmt in params:
            ws.cell(row=row, column=1, value=label).font = _BODY_FONT
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = _BODY_FONT
            vc.alignment = _ALIGN_RIGHT
            if fmt:
                vc.number_format = fmt
            row += 1
        row += 1  # blank row between sections

    _set_col_widths(ws, {"A": 30, "B": 20})
    ws.freeze_panes = "A2"


# ═══════════════════════════ PUBLIC API ════════════════════════════════

def generate_excel(inputs: dict, results: dict) -> bytes:
    """Build the complete workbook and return its bytes."""
    wb = Workbook()

    _build_executive_summary(wb, inputs, results)
    _build_cohort_matrix(wb, inputs, results)
    _build_revenue(wb, inputs, results)
    _build_costs(wb, inputs, results)
    _build_assumptions(wb, inputs, results)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
